import asyncio
import logging
import os
from datetime import datetime

from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery,
    Contact,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
import openpyxl
from openpyxl import Workbook

# ─────────────────────────────────────────────
#  SOZLAMALAR  –  bu yerlarni o'zgartiring
# ─────────────────────────────────────────────
BOT_TOKEN   = "8724145462:AAGmx4f6ag8LT8KK7jetWNwPr2U7tHiw0ow"          # @BotFather dan olingan token
ADMIN_ID    = 912613936                       # Sizning Telegram ID raqamingiz
CHANNELS    = ["@ekounion"]                   # Obuna tekshiriladigan kanallar
XLSX_FILE   = "users.xlsx"                    # Ma'lumotlar saqlanadigan fayl
# ─────────────────────────────────────────────

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = Router()


# ══════════════════════════════════════════════
#  FSM holatlari
# ══════════════════════════════════════════════
class Registration(StatesGroup):
    waiting_subscription = State()   # obuna tekshiruvi
    waiting_name         = State()
    waiting_region       = State()
    waiting_workplace    = State()
    waiting_birthday     = State()
    waiting_phone        = State()


# ══════════════════════════════════════════════
#  XLSX yordamchi funksiyalari
# ══════════════════════════════════════════════
HEADERS = ["Telegram ID", "Ism", "Viloyat/Tuman",
           "O'qish/Ish joyi", "Tug'ilgan kun", "Telefon", "Sana"]


def _ensure_workbook() -> openpyxl.Workbook:
    """Fayl mavjud bo'lsa ochadi, yo'q bo'lsa yangi yaratadi."""
    if os.path.exists(XLSX_FILE):
        return openpyxl.load_workbook(XLSX_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    ws.append(HEADERS)
    # Sarlavha ustunlarini kengaytirish
    for col in ws.iter_cols(1, len(HEADERS), 1, 1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
    wb.save(XLSX_FILE)
    return wb


def save_user(data: dict) -> None:
    """Bitta foydalanuvchini xlsx ga yozadi (takror bo'lsa yangilaydi)."""
    wb = _ensure_workbook()
    ws = wb.active

    # Mavjud qatorni izlaymiz (Telegram ID bo'yicha)
    existing_row = None
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(data["telegram_id"]):
            existing_row = row
            break

    row_data = [
        data["telegram_id"],
        data.get("name", ""),
        data.get("region", ""),
        data.get("workplace", ""),
        data.get("birthday", ""),
        data.get("phone", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]

    if existing_row:
        for i, cell in enumerate(existing_row):
            cell.value = row_data[i]
    else:
        ws.append(row_data)

    # Ustun enini avtomatik sozlash
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(XLSX_FILE)


def get_all_users() -> list[dict]:
    """Barcha foydalanuvchilarni ro'yxat ko'rinishida qaytaradi."""
    if not os.path.exists(XLSX_FILE):
        return []
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            users.append(dict(zip(HEADERS, row)))
    return users


# ══════════════════════════════════════════════
#  Obuna tekshirish
# ══════════════════════════════════════════════
async def check_subscriptions(bot: Bot, user_id: int) -> list[str]:
    """Foydalanuvchi obuna bo'lmagan kanallar ro'yxatini qaytaradi."""
    not_subscribed = []
    SUBSCRIBED_STATUSES = {"member", "creator", "administrator"}
    for channel in CHANNELS:
        try:
            member = await bot.get_chat_member(channel, user_id)
            if member.status not in SUBSCRIBED_STATUSES:
                not_subscribed.append(channel)
        except Exception as e:
            logger.warning("Kanal tekshirishda xato (%s): %s", channel, e)
            not_subscribed.append(channel)
    return not_subscribed


def subscription_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(text=f"📢 {ch}", url=f"https://t.me/{ch.lstrip('@')}")]
        for ch in CHANNELS
    ]
    buttons.append(
        [InlineKeyboardButton(text="✅ Obunani tekshirish", callback_data="check_sub")]
    )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def phone_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Telefon raqamimni yuborish", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


# ══════════════════════════════════════════════
#  Handlerlar
# ══════════════════════════════════════════════

@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    not_subscribed = await check_subscriptions(message.bot, message.from_user.id)

    if not_subscribed:
        channels_text = "\n".join(f"• {ch}" for ch in not_subscribed)
        await message.answer(
            f"👋 Xush kelibsiz!\n\n"
            f"Botdan foydalanish uchun avval quyidagi kanallarga obuna bo'ling:\n"
            f"{channels_text}\n\n"
            f"Obuna bo'lgandan so'ng <b>«✅ Obunani tekshirish»</b> tugmasini bosing.",
            parse_mode="HTML",
            reply_markup=subscription_keyboard(),
        )
        await state.set_state(Registration.waiting_subscription)
    else:
        await _ask_name(message, state)


@router.callback_query(F.data == "check_sub", StateFilter(Registration.waiting_subscription))
async def on_check_sub(callback: CallbackQuery, state: FSMContext) -> None:
    not_subscribed = await check_subscriptions(callback.bot, callback.from_user.id)

    if not_subscribed:
        channels_text = "\n".join(f"• {ch}" for ch in not_subscribed)
        await callback.answer("❌ Hali obuna bo'lmadingiz!", show_alert=True)
        await callback.message.edit_text(
            f"⚠️ Siz hali quyidagi kanallarga obuna bo'lmagansiz:\n{channels_text}\n\n"
            f"Obuna bo'lgach qayta tekshiring.",
            reply_markup=subscription_keyboard(),
        )
    else:
        await callback.answer("✅ Rahmat, obuna tasdiqlandi!")
        await callback.message.delete()
        await _ask_name(callback.message, state, from_callback=True, user=callback.from_user)


async def _ask_name(message: Message, state: FSMContext,
                    from_callback: bool = False, user=None) -> None:
    name = (user or message.from_user).first_name
    target = message if not from_callback else message

    await target.answer(
        f"👋 Salom, <b>{name}</b>!\n\n"
        f"Ro'yxatdan o'tish uchun bir necha savol beramiz.\n\n"
        f"1️⃣ <b>To'liq ismingizni</b> kiriting (Familiya Ism Sharif):",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Registration.waiting_name)


@router.message(StateFilter(Registration.waiting_name))
async def process_name(message: Message, state: FSMContext) -> None:
    name = message.text.strip()
    if len(name) < 3:
        await message.answer("❗ Iltimos, to'liq ismingizni kiriting (kamida 3 ta belgi).")
        return
    await state.update_data(name=name)
    await message.answer("2️⃣ <b>Viloyat / tumaningizni</b> kiriting:", parse_mode="HTML")
    await state.set_state(Registration.waiting_region)


@router.message(StateFilter(Registration.waiting_region))
async def process_region(message: Message, state: FSMContext) -> None:
    region = message.text.strip()
    if len(region) < 2:
        await message.answer("❗ Iltimos, to'g'ri viloyat yoki tuman nomini kiriting.")
        return
    await state.update_data(region=region)
    await message.answer(
        "3️⃣ <b>O'qish / ish joyingizni</b> kiriting\n"
        "<i>(masalan: ToshDU 2-kurs, Texnologiya MChJ)</i>:",
        parse_mode="HTML",
    )
    await state.set_state(Registration.waiting_workplace)


@router.message(StateFilter(Registration.waiting_workplace))
async def process_workplace(message: Message, state: FSMContext) -> None:
    workplace = message.text.strip()
    if len(workplace) < 2:
        await message.answer("❗ Iltimos, o'qish yoki ish joyingizni kiriting.")
        return
    await state.update_data(workplace=workplace)
    await message.answer(
        "4️⃣ <b>Tug'ilgan kuningizni</b> kiriting\n"
        "<i>Format: kun.oy.yil — masalan: 15.03.2000</i>:",
        parse_mode="HTML",
    )
    await state.set_state(Registration.waiting_birthday)


@router.message(StateFilter(Registration.waiting_birthday))
async def process_birthday(message: Message, state: FSMContext) -> None:
    raw = message.text.strip()
    try:
        dt = datetime.strptime(raw, "%d.%m.%Y")
        # Yoshni tekshiramiz: 5–100 yosh
        age = (datetime.now() - dt).days // 365
        if not (5 <= age <= 100):
            raise ValueError("Noto'g'ri yosh")
        birthday = dt.strftime("%d.%m.%Y")
    except ValueError:
        await message.answer(
            "❗ Sana noto'g'ri kiritildi.\n"
            "Iltimos, <b>kun.oy.yil</b> formatida kiriting, masalan: <code>15.03.2000</code>",
            parse_mode="HTML",
        )
        return

    await state.update_data(birthday=birthday)
    await message.answer(
        "5️⃣ <b>Telefon raqamingizni yuboring</b>\n"
        "Quyidagi tugmani bosing yoki raqamni qo'lda kiriting (+998XXXXXXXXX):",
        parse_mode="HTML",
        reply_markup=phone_keyboard(),
    )
    await state.set_state(Registration.waiting_phone)


@router.message(StateFilter(Registration.waiting_phone), F.contact)
async def process_phone_contact(message: Message, state: FSMContext) -> None:
    contact: Contact = message.contact
    phone = contact.phone_number
    if not phone.startswith("+"):
        phone = "+" + phone
    await _finish_registration(message, state, phone)


@router.message(StateFilter(Registration.waiting_phone), F.text)
async def process_phone_text(message: Message, state: FSMContext) -> None:
    phone = message.text.strip().replace(" ", "").replace("-", "")
    if not (phone.startswith("+") and 10 <= len(phone) <= 15 and phone[1:].isdigit()):
        await message.answer(
            "❗ Noto'g'ri format. Iltimos, raqamni <code>+998901234567</code> ko'rinishida kiriting "
            "yoki tugma orqali yuboring.",
            parse_mode="HTML",
        )
        return
    await _finish_registration(message, state, phone)


async def _finish_registration(message: Message, state: FSMContext, phone: str) -> None:
    data = await state.get_data()
    data["phone"] = phone
    data["telegram_id"] = message.from_user.id

    save_user(data)
    await state.clear()

    summary = (
        "✅ <b>Ro'yxatdan o'tish muvaffaqiyatli yakunlandi!</b>\n\n"
        f"👤 Ism: <b>{data['name']}</b>\n"
        f"🏙 Viloyat/tuman: <b>{data['region']}</b>\n"
        f"🏫 O'qish/ish joyi: <b>{data['workplace']}</b>\n"
        f"🎂 Tug'ilgan kun: <b>{data['birthday']}</b>\n"
        f"📱 Telefon: <b>{phone}</b>\n\n"
        "Ma'lumotlaringiz saqlandi. Rahmat! 🙏"
    )
    await message.answer(summary, parse_mode="HTML", reply_markup=ReplyKeyboardRemove())


# ══════════════════════════════════════════════
#  Admin komandalar
# ══════════════════════════════════════════════

@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Bu komanda faqat admin uchun.")
        return

    users = get_all_users()
    if not users:
        await message.answer("📭 Hozircha hech qanday foydalanuvchi ro'yxatdan o'tmagan.")
        return

    # Statistika xabari
    await message.answer(
        f"📊 <b>Jami foydalanuvchilar: {len(users)} ta</b>\n"
        f"📁 Fayl yuklanmoqda...",
        parse_mode="HTML",
    )

    # Fayl yuborish
    if os.path.exists(XLSX_FILE):
        from aiogram.types import FSInputFile
        file = FSInputFile(XLSX_FILE, filename="users_export.xlsx")
        await message.answer_document(
            file,
            caption=f"📋 Barcha foydalanuvchilar ro'yxati ({len(users)} ta)\n"
                    f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )
    else:
        await message.answer("❌ Fayl topilmadi.")


@router.message(Command("stats"))
async def cmd_stats(message: Message) -> None:
    if message.from_user.id != ADMIN_ID:
        await message.answer("⛔ Bu komanda faqat admin uchun.")
        return

    users = get_all_users()
    if not users:
        await message.answer("📭 Foydalanuvchilar yo'q.")
        return

    # Viloyatlar bo'yicha statistika
    regions: dict[str, int] = {}
    for u in users:
        r = str(u.get("Viloyat/Tuman") or "Noma'lum")
        regions[r] = regions.get(r, 0) + 1

    top_regions = sorted(regions.items(), key=lambda x: x[1], reverse=True)[:10]
    region_text = "\n".join(f"  • {r}: {c} ta" for r, c in top_regions)

    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Jami foydalanuvchilar: <b>{len(users)}</b>\n\n"
        f"🏙 Viloyatlar bo'yicha (Top-10):\n{region_text}",
        parse_mode="HTML",
    )


@router.message(Command("help"))
async def cmd_help(message: Message) -> None:
    if message.from_user.id == ADMIN_ID:
        text = (
            "🤖 <b>Bot komandalar (Admin)</b>\n\n"
            "/start — Botni qayta ishga tushirish\n"
            "/export — Barcha foydalanuvchilarni xlsx sifatida yuklab olish\n"
            "/stats — Statistika ko'rish\n"
            "/help — Yordam\n"
        )
    else:
        text = (
            "🤖 <b>Bot komandalar</b>\n\n"
            "/start — Ro'yxatdan o'tish\n"
            "/help — Yordam\n"
        )
    await message.answer(text, parse_mode="HTML")


# ══════════════════════════════════════════════
#  Ishga tushirish
# ══════════════════════════════════════════════
async def main() -> None:
    bot = Bot(token=BOT_TOKEN)
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)

    logger.info("Bot ishga tushdi ✅")
    await dp.start_polling(bot, allowed_updates=["message", "callback_query"])


if __name__ == "__main__":
    asyncio.run(main())